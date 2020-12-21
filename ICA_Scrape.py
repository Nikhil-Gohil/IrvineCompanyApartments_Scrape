##Script to scrape apartment cost from Irvine Company apartments website
from bs4 import BeautifulSoup
from selenium import webdriver
import requests
import csv
import sys
import pprint
import re
import openpyxl
import datetime
import time

printer = pprint.PrettyPrinter(indent=4)


def main():
    wb = open_excel_workbook()
    outfile = build_outfile_name()
    url_dict = create_populate_url_dict()

    for name, url in url_dict.items():
        print("Processing {}".format(name))
        parse_ica_page(name, url, wb)

    wb.save(outfile)


def open_excel_workbook():
    """
    Create an openpyxl workbook, delete default sheet and return it.
    """
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    # wb.remove_s(wb.get_sheet_by_name("Sheet"))
    return wb


def build_outfile_name():
    """
    Builds output filename based on current date and time.
    """
    now = datetime.datetime.now()
    str1 = now.strftime("%a_%b%d_%H_%M")
    fn = "Apartment_Availability_" + str1 + ".xlsx"
    return fn


def create_populate_url_dict():
    """
    Create a dict and populate it with name - url pairs for the different Irvine Company Apartment locations.
    """
    url_dict = {}
    url_dict[
        "The Village"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/village-at-irvine-spectrum/availability.html"
    url_dict[
        "Oak Glen"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/oak-creek/oak-glen/availability.html"
    url_dict[
        "Cypress Village"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/cypress-village/communities/availability.html"
    url_dict[
        "Avella"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/cypress-village/avella/availability.html"
    url_dict[
        "Quail Hill"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/quail-hill/communities/availability.html"
    url_dict[
        "The Park"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/the-park/availability.html"
    url_dict[
        "Woodbury Court"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/woodbury/woodbury/woodbury-court/availability.html"
    url_dict[
        "Centerpointe"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/centerpointe/availability.html"
    url_dict[
        "Westview"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/westview/availability.html"
    url_dict[
        "Los Olivos"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/los-olivos/availability.html"
    url_dict[
        "Promenade"
    ] = "https://www.irvinecompanyapartments.com/locations/orange-county/irvine/spectrum/promenade-at-spectrum/availability.html"

    return url_dict


def create_populate_sheet_headers(name, wb):
    """
    Create a new sheet in the workbook with the name of the Apartment complex. Also populates the headers.
    """
    wb.create_sheet(title=name)
    sheet = wb[name]
    array = [
        "Identifier",
        "Building No.",
        "Building Name",
        "Floorplan",
        "Amenity",
        "Beds",
        "Baths",
        "Size (Sq. Ft.)",
        "Price ($)",
        "Term (months)",
        "Date Available",
    ]
    for index, val in enumerate(array):
        sheet.cell(row=1, column=index + 1).value = val
    return wb[name]


def convert_to_int(str1):
    """
    Try to convert a string to int if it is numeric, otherwise returns the string itself.
    """
    try:
        num = int(str1)
        return num
    except ValueError:
        return str1


def convert_to_float(str1):
    """
    Try to convert a string to float if it is numeric, otherwise returns the string itself.
    """
    try:
        num = float(str1)
        return num
    except ValueError:
        return str1


def get_page_source(url):
    """
    Open browser using webdriver, load the url and get the html source.
    """
    browser = webdriver.Chrome()
    browser.get(url)
    time.sleep(10)
    html_source = browser.page_source
    browser.quit()
    return html_source


def parse_ica_page(name, url, wb):
    """
    Parses the html source of the availablility page of the apartment complex website and writes all the currently available apartments to an exccel sheet in the workbook provided.
    """
    html_source = get_page_source(url)
    soup = BeautifulSoup(html_source, "lxml")
    # the "results-list loaded" list has all the currently available apartments in an html list.
    all_apartments_list = soup.find_all("ul", class_="results-list loaded")
    # then we can find all list entries, each entry is an apartment complex.
    individual_apartments = all_apartments_list[0].find_all("li")

    sheet = create_populate_sheet_headers(name, wb)
    row = 1
    column = 1

    # now we iterate over every available apartment, parse the html and write data to excel sheet.
    for apartment in individual_apartments:
        text = apartment.get_text()
        if re.match("\n\nNeed More Options.*", text):
            continue

        name = apartment.find("h5").get_text()

        floorplan_details = name.split(" - ")
        try:
            identifier, building, floorplan = floorplan_details
        except ValueError:
            identifier, building = floorplan_details
            floorplan = "-"
        building_array = building.split(" ")
        building_number = building_array.pop(0)
        building_name = " ".join(building_array)

        amenity = apartment.find("div", class_="featured-amenity").get_text()
        if not amenity:
            amenity = "None"

        divs = apartment.find_all("div")

        details = divs[1].get_text()
        detail_array = details.split(" / ")
        no_of_beds, no_of_baths, size = detail_array

        no_of_beds = re.sub(" Bed$", "", no_of_beds)
        no_of_baths = re.sub(" Bath$", "", no_of_baths)
        size = re.sub(" Sq\. Ft\.$", "", size)
        size = re.sub(",", "", size)

        pricing = divs[2].get_text()
        price_array = pricing.split(" / ")
        cost, term_length = price_array
        cost = re.sub("^\$", "", cost)
        cost = re.sub(",", "", cost)
        term_length = re.sub(" Months$", "", term_length)

        availability = divs[3].get_text()
        availability = re.sub("^Available ", "", availability)

        no_of_beds = convert_to_int(no_of_beds)
        no_of_baths = convert_to_float(no_of_baths)
        size = convert_to_int(size)
        cost = convert_to_int(cost)
        term_length = convert_to_int(term_length)
        building_number = convert_to_int(building_number)

        array = [
            identifier,
            building_number,
            building_name,
            floorplan,
            amenity,
            no_of_beds,
            no_of_baths,
            size,
            cost,
            term_length,
            availability,
        ]
        row += 1
        for index, val in enumerate(array):
            col = index + 1
            sheet.cell(row=row, column=index + 1).value = val

    # adding filters to headers
    sheet.auto_filter.ref = sheet.dimensions

    # changing cell width for better looking sheets
    dims = {}
    for row in sheet.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 3


if __name__ == "__main__":
    main()